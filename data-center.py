import sys, os, re, time, platform, subprocess, smtplib, schedule
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ─────────────────────────────────────────────
#  CONFIGURATION  ← Edit these values
# ─────────────────────────────────────────────
INPUT_EXCEL    = "our excel file name  "
GMAIL_SENDER   = 
GMAIL_PASSWORD = 
MANAGER_EMAIL = [
    
]
REPORT_FOLDER  = "reports"
PING_COUNT     = 3
HIGH_LAT_MS    = 100
SKIP_SHEETS    = {"SNMP"}
HISTORY_FILE   = "uptime_history.xlsx"
CHECK_INTERVAL_HOURS = 3

# ─────────────────────────────────────────────

OS = platform.system().lower()

# ══════════════════════════════════════════════
#  PREMIUM BRANDED CORPORATE PALETTE (MUTED)
# ══════════════════════════════════════════════
COLOR_PRIMARY_NAVY = "0A192F"  # Deep luxury navy blue for primary titles
COLOR_HEADER_SLATE = "1E293B"  # Professional slate/charcoal for table headers
COLOR_KPI_BLUE     = "E2F1FF"  # Soft icy blue for structural data tiles
COLOR_KPI_TXT_BLUE = "0F4C81"  

COLOR_UP_GREEN     = "E6F4EA"  # Soft mint green background
COLOR_UP_TXT       = "137333"  # Clear legible forest green text

COLOR_DOWN_RED     = "FCE8E6"  # Soft rose/coral background 
COLOR_DOWN_TXT     = "C5221F"  # Clear legible ruby/crimson text

COLOR_UNASG_GREY   = "F1F3F4"  # Light neutral silver gray
COLOR_UNASG_TXT    = "5F6368"  # Dark gray text

COLOR_WARN_YLLW    = "FEF7E0"  # Soft sand cream background
COLOR_WARN_TXT     = "B06000"  # Soft brown-gold warning text

# ══════════════════════════════════════════════
#  STYLE HELPERS
# ══════════════════════════════════════════════
def _bdr(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val, bold=False, size=9, fc="000000", bg=None,
          halign="center", valign="center", wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Segoe UI", bold=bold, size=size, color=fc)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = _bdr()
    return c

def _merge(ws, r1, c1, r2, c2, val, bold=False, size=9, fc="000000",
           bg=None, halign="center", valign="center", wrap=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=val)
    c.font = Font(name="Segoe UI", bold=bold, size=size, color=fc)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        for r in range(r1, r2 + 1):
            for col_idx in range(c1, c2 + 1):
                ws.cell(row=r, column=col_idx).fill = PatternFill("solid", start_color=bg)
                ws.cell(row=r, column=col_idx).border = _bdr()
    c.border = _bdr()
    return c

def kpi_tile(ws, r1, c1, r2, c2, label, value, pct, fc, bg):
    _merge(ws, r1,   c1, r1,   c2, label, bold=True, size=8,  fc=fc, bg=bg)
    _merge(ws, r1+1, c1, r2-1, c2, str(value), bold=True, size=18, fc=fc, bg=bg)
    if pct is not None:
        _merge(ws, r2, c1, r2, c2, pct, bold=True, size=8, fc=fc, bg=bg)
    else:
        _merge(ws, r2, c1, r2, c2, "", bg=bg)

def health_grade(avail_pct):
    if avail_pct >= 95: return "A", "EXCELLENT STATUS", COLOR_UP_TXT, COLOR_UP_GREEN
    if avail_pct >= 85: return "B", "MOSTLY OPERATIONAL", COLOR_KPI_TXT_BLUE, COLOR_KPI_BLUE
    if avail_pct >= 75: return "C", "ATTENTION REQUIRED", COLOR_WARN_TXT, COLOR_WARN_YLLW
    return                     "D", "CRITICAL INCIDENT", COLOR_DOWN_TXT, COLOR_DOWN_RED

def ms_color(ms):
    if ms is None:      return COLOR_DOWN_RED, COLOR_DOWN_TXT
    if ms <= 50:        return COLOR_UP_GREEN, COLOR_UP_TXT
    if ms <= 100:       return COLOR_WARN_YLLW, COLOR_WARN_TXT
    return                     COLOR_DOWN_RED, COLOR_DOWN_TXT

# ══════════════════════════════════════════════
#  IP EXTRACTION
# ══════════════════════════════════════════════
# ══════════════════════════════════════════════
#  IP EXTRACTION (SMART COLUMN MAPPING)
# ══════════════════════════════════════════════
IP_PAT = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')

def extract_sheet_ips(filepath):
    wb   = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    data = {}
    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue
        ws   = wb[sname]
        rows = []
        seen = set()
        for row in ws.iter_rows(values_only=True):
            # Clean up cells to string format
            row = [str(c).strip() if c else "" for c in row]
            
            # 1. Locate the IP address in the row
            ip_col_idx = -1
            for i, cell in enumerate(row):
                if IP_PAT.match(cell):
                    ip_col_idx = i
                    break
                    
            if ip_col_idx != -1:
                ip_addr = row[ip_col_idx]
                if ip_addr in seen:
                    continue  # Skip duplicate IPs
                seen.add(ip_addr)
                
                # 2. Extract VLAN: The first digit-only string strictly AFTER the IP address
                vlan = next((row[j] for j in range(ip_col_idx + 1, len(row)) if row[j] and row[j].isdigit()), "")
                
                # 3. Gather all other valid text fields (excluding IP and VLAN)
                text_fields = []
                vlan_skipped = False
                for j, cell in enumerate(row):
                    if j == ip_col_idx:
                        continue
                    # Skip the first occurrence of the VLAN after the IP to avoid adding it to text
                    if j > ip_col_idx and cell == vlan and not vlan_skipped:
                        vlan_skipped = True
                        continue
                    # Grab valid text that isn't an IP and isn't empty
                    if cell and not IP_PAT.match(cell):
                        text_fields.append(cell)
                
                # 4. Dynamically assign fields based on the text found
                device = "N/A"
                hostname = "N/A"
                location = ""
                
                if ip_col_idx == 0:
                    # If IP is the first column (e.g., VM'S sheet): [IP Address, Hostname, OS/Location]
                    if len(text_fields) >= 1:
                        hostname = text_fields[0]
                    if len(text_fields) >= 2:
                        location = " | ".join(text_fields[1:])
                else:
                    # If IP is further right (Most Sheets): [Device, Hostname, Location/OS, IP Address, VLAN]
                    if len(text_fields) >= 1:
                        device = text_fields[0]
                    if len(text_fields) >= 2:
                        hostname = text_fields[1]
                    if len(text_fields) >= 3:
                        location = " | ".join(text_fields[2:])
                
                # Add to row list
                rows.append((device, hostname, ip_addr, vlan, location))
                
        if rows:
            data[sname] = rows
    wb.close()
    return data

# ══════════════════════════════════════════════
#  PING ENGINE
# ══════════════════════════════════════════════
def ping(ip):
    cmd = (["ping","-n",str(PING_COUNT),"-w","1000",ip] if OS=="windows" else ["ping","-c",str(PING_COUNT),"-W","1",ip])
    try:
        r  = subprocess.run(cmd, capture_output=True, text=True, timeout=25)
        lo = r.stdout.lower()
        received = 0
        if OS == "windows":
            m = re.search(r'received\s*=\s*(\d+)', lo)
            if m: received = int(m.group(1))
        else:
            m = re.search(r'(\d+) received', lo)
            if m: received = int(m.group(1))

        if received > 0:
            avg = None
            for line in r.stdout.splitlines():
                ll = line.lower()
                if "average" in ll:
                    parts = line.replace("="," ").replace(","," ").split()
                    for k,p in enumerate(parts):
                        if "average" in p.lower() and k+1<len(parts):
                            try: avg=float(parts[k+1].replace("ms","").strip())
                            except: pass
                elif "min/avg/max" in ll:
                    m2 = re.search(r'=([\d.]+)/([\d.]+)/([\d.]+)',line)
                    if m2:
                        try: avg=float(m2.group(2))
                        except: pass
            last = next((l for l in r.stdout.splitlines() if "reply from" in l.lower() or "bytes from" in l.lower()), "Reply received.")
            return "UP", avg, last
        elif "unreachable" in lo:
            return "UNASSIGNED", None, "Destination host unreachable."
        else:
            return "DOWN", None, "Request timed out."
    except subprocess.TimeoutExpired:
        return "DOWN", None, "Request timed out."
    except Exception as e:
        return "DOWN", None, f"Error: {e}"

# ══════════════════════════════════════════════
#  EXECUTIVE SUMMARY SHEET (RE-ARCHITECTED)
# ══════════════════════════════════════════════
def build_executive_summary(ws, all_results, now, scan_duration):
    ws.sheet_view.showGridLines = False
    today  = now.strftime("%d-%b-%Y")
    ttime  = now.strftime("%I:%M %p")

    for i, w in enumerate([3, 24, 15, 18, 18, 15, 22, 28, 3], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── PRIMARY BUSINESS BANNER ──
    ws.row_dimensions[1].height = 42
    _merge(ws, 1, 2, 1, 8, "SNOC INFRASTRUCTURE REPORT  –  EXECUTIVE HEALTH BRIEF", bold=True, size=13, fc="FFFFFF", bg=COLOR_PRIMARY_NAVY)

    ws.row_dimensions[2].height = 4
    ws.row_dimensions[3].height = 20
    _merge(ws, 3, 2, 3, 8, f"Timeline: {today} at {ttime}  |  Operational window: 2-Hour Automated Sync  |  Cycle Span: {scan_duration}", size=8.5, fc="FFFFFF", bg=COLOR_HEADER_SLATE)

    # Calculate metrics
    grand_t = grand_u = grand_d = grand_un = 0
    for sname, results in all_results.items():
        grand_t  += len(results)
        grand_u  += sum(1 for r in results if r[5]=="UP")
        grand_d  += sum(1 for r in results if r[5]=="DOWN")
        grand_un += sum(1 for r in results if r[5]=="UNASSIGNED")
    avail = (grand_u / grand_t * 100) if grand_t else 0
    grade, grade_msg, grade_fc, grade_bg = health_grade(avail)

    ws.row_dimensions[4].height = 10
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 24
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 12

    # Score block
    _merge(ws, 5, 2, 5, 3, "NET SYSTEM GRADE", bold=True, size=8, fc=grade_fc, bg=grade_bg)
    _merge(ws, 6, 2, 6, 3, f"  {grade}  ", bold=True, size=28, fc=grade_fc, bg=grade_bg)
    _merge(ws, 7, 2, 7, 3, grade_msg, bold=True, size=8, fc=grade_fc, bg=grade_bg)

    # Executive Summary Blocks
    kpi_tile(ws, 5, 4, 7, 4, "TOTAL NETWORK NODE IPS", grand_t, None, "333333", "F8F9FA")
    kpi_tile(ws, 5, 5, 7, 5, "ACTIVE CHANNELS (UP)", grand_u, f"{avail:.2f}% Availability", COLOR_UP_TXT, COLOR_UP_GREEN)
    kpi_tile(ws, 5, 6, 7, 6, "CRITICAL INCIDENTS (DOWN)", grand_d, f"{grand_d} Active Alerts" if grand_d > 0 else "All Clear", COLOR_DOWN_TXT if grand_d > 0 else COLOR_UP_TXT, COLOR_DOWN_RED if grand_d > 0 else COLOR_UP_GREEN)
    kpi_tile(ws, 5, 7, 7, 7, "UNALLOCATED NET BLOCKS", grand_un, "Isolated Subnets", COLOR_UNASG_TXT, COLOR_UNASG_GREY)
    kpi_tile(ws, 5, 8, 7, 8, "GLOBAL AGGREGATE UPTIME", f"{avail:.2f}%", "Target SLA: 99.90%", COLOR_KPI_TXT_BLUE, COLOR_KPI_BLUE)

    # ── CATEGORY MATRIX SUMMARY ──
    ws.row_dimensions[9].height  = 6
    ws.row_dimensions[10].height = 22
    _merge(ws, 10, 2, 10, 8, " SYSTEM DOMAIN CLASSIFICATION STATUS", bold=True, size=10, fc="FFFFFF", bg=COLOR_PRIMARY_NAVY, halign="left")

    ws.row_dimensions[11].height = 22
    for ci, h in zip(range(2, 9), ["Infrastructure Domain Group", "Total Nodes", "Online (UP)", "Alerting (DOWN)", "Unassigned Blocks", "SLA Performance %", "Domain Index"]):
        _cell(ws, 11, ci, h, bold=True, size=9, fc="FFFFFF", bg=COLOR_HEADER_SLATE)

    current_row = 12
    for idx, (sname, results) in enumerate(all_results.items(), 12):
        t  = len(results)
        u  = sum(1 for r in results if r[5]=="UP")
        d  = sum(1 for r in results if r[5]=="DOWN")
        un = sum(1 for r in results if r[5]=="UNASSIGNED")
        av = (u / t * 100) if t else 0
        gr, _, gfc, gbg = health_grade(av)
        bg2 = "F8F9FA" if idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[current_row].height = 20

        _cell(ws, current_row, 2, sname, size=9, bg=bg2, halign="left", bold=True)
        _cell(ws, current_row, 3, t,    size=9, bg=bg2)
        _cell(ws, current_row, 4, u,    size=9, fc=COLOR_UP_TXT, bg=bg2, bold=True)
        _cell(ws, current_row, 5, d,    size=9, fc=COLOR_DOWN_TXT if d > 0 else COLOR_UP_TXT, bg=COLOR_DOWN_RED if d > 0 else bg2, bold=d > 0)
        _cell(ws, current_row, 6, un,   size=9, fc=COLOR_UNASG_TXT, bg=bg2)
        _cell(ws, current_row, 7, f"{av:.2f}%", size=9, fc=COLOR_UP_TXT if av >= 90 else COLOR_DOWN_TXT, bg=bg2, bold=True)
        _cell(ws, current_row, 8, f"Class {gr}", size=9, fc=gfc, bg=gbg, bold=True)
        current_row += 1

    # ── IMPACT-BASED CRITICAL OUTAGES MATRIX ──
    current_row += 2
    ws.row_dimensions[current_row].height = 24
    
    all_down = []
    for sname, results in all_results.items():
        for r in results:
            if r[5] == "DOWN":
                all_down.append((sname, r))

    _merge(ws, current_row, 2, current_row, 8, f" RISK THREAT BRIEF: STRATIFIED LIVE ALERTS ({len(all_down)} Total Critical Faults)", bold=True, size=10, fc="FFFFFF", bg=COLOR_DOWN_TXT if all_down else COLOR_UP_TXT, halign="left")
    
    current_row += 1
    if all_down:
        # Group dynamic arrays into Tier impacts for management analysis
        core_infra_sheets = {"networking devices&physical-srs", "servers nics", "netapp-storage"}
        vm_storage_sheets = {"vm's"}
        
        core_impact = []
        vm_impact = []
        user_facility_impact = []
        
        for item in all_down:
            sname_lower = item[0].lower()
            if sname_lower in core_infra_sheets:
                core_impact.append(item)
            elif sname_lower in vm_storage_sheets:
                vm_impact.append(item)
            else:
                user_facility_impact.append(item)
                
        impact_tiers = [
            ("TIER 1: CORE INFRASTRUCTURE (HIGH THREAT OUTAGE)", core_impact, "7B1FA2"),
            ("TIER 2: ENGINE VIRTUALIZATION & DATA LAYERS (OPERATIONAL IMPACT)", vm_impact, "1976D2"),
            ("TIER 3: END-USER CLUSTERS & PERIPHERAL ACCESS", user_facility_impact, "E65100")
        ]
        
        for tier_title, items_list, tier_color in impact_tiers:
            if not items_list:
                continue
            
            ws.row_dimensions[current_row].height = 18
            _merge(ws, current_row, 2, current_row, 8, f"  {tier_title}", bold=True, size=8.5, fc="FFFFFF", bg=tier_color, halign="left")
            current_row += 1
            
            ws.row_dimensions[current_row].height = 18
            for ci, h in zip(range(2, 9), ["Resource Origin", "Target Host IP", "Equipment Flag", "System Name", "Physical Zone/VLAN", "Incident Log", "Ownership Vector"]):
                _cell(ws, current_row, ci, h, bold=True, size=8, fc="FFFFFF", bg=COLOR_HEADER_SLATE)
            current_row += 1
            
            for sname, (device, hostname, ip, vlan, loc, status, avg_ms, last) in items_list:
                ws.row_dimensions[current_row].height = 18
                _cell(ws, current_row, 2, sname, size=8, bg="FAFAFA", halign="left")
                _cell(ws, current_row, 3, ip, size=8, bg="FAFAFA", bold=True, fc=COLOR_DOWN_TXT)
                _cell(ws, current_row, 4, device, size=8, bg="FAFAFA", halign="left")
                _cell(ws, current_row, 5, hostname, size=8, bg="FAFAFA", halign="left")
                _cell(ws, current_row, 6, f"VLAN {vlan if vlan else 'N/A'}", size=8, bg="FAFAFA")
                _cell(ws, current_row, 7, last, size=8, bg="FAFAFA", halign="left", wrap=True)
                _cell(ws, current_row, 8, "Internal SNOC Action Team", size=8, bg="FAFAFA", bold=True, fc=COLOR_DOWN_TXT)
                current_row += 1
    else:
        ws.row_dimensions[current_row].height = 20
        _merge(ws, current_row, 2, current_row, 8, "  All monitoring lines healthy. Zero outages detected across infrastructure zones.", size=9, fc=COLOR_UP_TXT, bg=COLOR_UP_GREEN, halign="left")
        current_row += 1

# ══════════════════════════════════════════════
#  MASTER DASHBOARD SHEET
# ══════════════════════════════════════════════
def build_dashboard(ws, all_results, now):
    ws.sheet_view.showGridLines = False
    today = now.strftime("%d-%b-%Y")
    ttime = now.strftime("%I:%M %p")

    for i, w in enumerate([3, 28, 14, 12, 12, 14, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 36
    _merge(ws, 2, 2, 2, 8, "SNOC NETWORK SYSTEMS INTERACTIVE DASHBOARD", bold=True, size=14, fc="FFFFFF", bg=COLOR_PRIMARY_NAVY)

    for rr, lbl, val in [(3, "Execution Date", today), (4, "Execution Time", ttime), (5, "Sync Matrix Rate", "Every 2 Hours"), (6, "Automation Node", "IP Monitoring Engine")]:
        ws.row_dimensions[rr].height = 18
        _cell(ws, rr, 2, lbl, bold=True, size=8.5, fc=COLOR_KPI_TXT_BLUE, bg=COLOR_KPI_BLUE, halign="left")
        _merge(ws, rr, 3, rr, 4, val, size=8.5, bg="FAFAFA", halign="left")

    ws.row_dimensions[7].height = 12
    ws.row_dimensions[8].height = 22
    _merge(ws, 8, 2, 8, 8, " DATA SEGMENT CLUSTER BREAKDOWN", bold=True, size=10, fc="FFFFFF", bg=COLOR_PRIMARY_NAVY, halign="left")

    ws.row_dimensions[9].height = 20
    for ci, h in zip(range(2, 9), ["Category", "Total Target IPs", "UP", "DOWN", "UNASSIGNED", "Availability Rate", "Subnet Index"]):
        _cell(ws, 9, ci, h, bold=True, size=9, fc="FFFFFF", bg=COLOR_HEADER_SLATE)

    grand = [0, 0, 0, 0]
    start_row = 10
    for idx, (sname, results) in enumerate(all_results.items(), 1):
        rn = 9 + idx
        ws.row_dimensions[rn].height = 18
        t  = len(results)
        u  = sum(1 for r in results if r[5]=="UP")
        d  = sum(1 for r in results if r[5]=="DOWN")
        un = sum(1 for r in results if r[5]=="UNASSIGNED")
        av = (u / t * 100) if t else 0
        gr, _, gfc, gbg = health_grade(av)
        bg2 = "F8F9FA" if idx % 2 else "FFFFFF"

        _cell(ws, rn, 2, sname, size=8.5, bg=bg2, halign="left", bold=True)
        _cell(ws, rn, 3, t,  size=8.5, bg=bg2)
        _cell(ws, rn, 4, u,  size=8.5, fc=COLOR_UP_TXT, bg=bg2, bold=True)
        _cell(ws, rn, 5, d,  size=8.5, fc=COLOR_DOWN_TXT if d > 0 else COLOR_UP_TXT, bg=COLOR_DOWN_RED if d > 0 else bg2, bold=d > 0)
        _cell(ws, rn, 6, un, size=8.5, fc=COLOR_UNASG_TXT, bg=bg2)
        _cell(ws, rn, 7, f"{av:.2f}%", size=8.5, fc=COLOR_UP_TXT if av >= 90 else COLOR_DOWN_TXT, bg=bg2, bold=True)
        _cell(ws, rn, 8, f"Class {gr}", size=8.5, fc=gfc, bg=gbg, bold=True)

        grand[0] += t; grand[1] += u; grand[2] += d; grand[3] += un

    tr = 9 + len(all_results) + 1
    ws.row_dimensions[tr].height = 22
    gv = (grand[1] / grand[0] * 100) if grand[0] else 0
    _cell(ws, tr, 2, "NET SYSTEM ACCUMULATION", bold=True, size=9, fc="FFFFFF", bg=COLOR_HEADER_SLATE, halign="left")
    _cell(ws, tr, 3, grand[0], bold=True, size=9, fc="FFFFFF", bg=COLOR_HEADER_SLATE)
    _cell(ws, tr, 4, grand[1], bold=True, size=9, fc=COLOR_UP_GREEN, bg=COLOR_HEADER_SLATE)
    _cell(ws, tr, 5, grand[2], bold=True, size=9, fc=COLOR_DOWN_RED if grand[2] > 0 else "FFFFFF", bg=COLOR_HEADER_SLATE)
    _cell(ws, tr, 6, grand[3], bold=True, size=9, fc="FFFFFF", bg=COLOR_HEADER_SLATE)
    _cell(ws, tr, 7, f"{gv:.2f}%", bold=True, size=9, fc=COLOR_UP_GREEN if gv >= 90 else COLOR_DOWN_RED, bg=COLOR_PRIMARY_NAVY)
    
    gr2, _, gfc2, gbg2 = health_grade(gv)
    _cell(ws, tr, 8, f"TOTAL {gr2}", bold=True, size=9, fc=gfc2, bg=gbg2)

# ══════════════════════════════════════════════
#  GRANULAR TELEMETRY WORKBOOK BUILDERS
# ══════════════════════════════════════════════
def build_sheet_report(ws, sheet_name, results, now):
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "B11"  # Natively lock column context blocks for smooth management scrolls
    today = now.strftime("%d-%b-%Y")
    ttime = now.strftime("%I:%M %p")

    # Metrics
    t  = len(results)
    u  = sum(1 for r in results if r[5]=="UP")
    d  = sum(1 for r in results if r[5]=="DOWN")
    un = sum(1 for r in results if r[5]=="UNASSIGNED")
    av = (u / t * 100) if t else 0
    grade, phrase, gfc, gbg = health_grade(av)

    # Base Matrix Grid Width Structs
    for i, w in enumerate([4, 6, 16, 26, 26, 10, 15, 12, 12, 12, 15, 42, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 36
    _merge(ws, 2, 2, 2, 12, f"TELEMETRY DATA MATRIX – INTERACTIVE REGION: {sheet_name.upper()}", bold=True, size=11, fc="FFFFFF", bg=COLOR_PRIMARY_NAVY)

    # Secondary Mini Score Headers
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18
    headers_lbls = ["HEALTH GRADE", "TOTAL BLOCK IPs", "NODES ONLINE", "ALERT LOGS", "UNMAPPED RESERVES", "AVAILABILITY"]
    vals = [grade, t, u, d, un, f"{av:.2f}%"]
    fcs = [gfc, "000000", COLOR_UP_TXT, COLOR_DOWN_TXT if d > 0 else "000000", COLOR_UNASG_TXT, COLOR_UP_TXT if av >= 90 else COLOR_DOWN_TXT]
    bgs = [gbg, "FAFAFA", COLOR_UP_GREEN, COLOR_DOWN_RED if d > 0 else "FAFAFA", COLOR_UNASG_GREY, gbg]

    for idx, (lbl, val, fc, bg) in enumerate(zip(headers_lbls, vals, fcs, bgs)):
        col_start = 2 + (idx * 2)
        _merge(ws, 4, col_start, 4, col_start + 1, lbl, bold=True, size=7.5, fc="555555", bg="EFEFEF")
        _merge(ws, 5, col_start, 5, col_start + 1, str(val), bold=True, size=10, fc=fc, bg=bg)

    ws.row_dimensions[7].height = 18
    _merge(ws, 7, 2, 7, 12, phrase, bold=True, size=9, fc=gfc, bg=gbg)

    ws.row_dimensions[9].height = 20
    _merge(ws, 9, 2, 9, 12, f"SYS ARCHITECTURE SNAPSHOT LOOP – LINE METRICS | Run Window: {today} {ttime}", bold=True, size=9, fc="FFFFFF", bg=COLOR_HEADER_SLATE, halign="left")

    ws.row_dimensions[10].height = 22
    hdrs = ["S.No", "IP Address", "Device Module", "Hostname System", "VLAN", "Physical Space", "Status Block", "Latency (ms)", "Packet Loss", "Timestamp","Daily uptime %","Daily Downtime", "Raw Diagnostic Feedback", "Remarks Context"]
    for ci, h in enumerate(hdrs, 2):
        _cell(ws, 10, ci, h, bold=True, size=8.5, fc="FFFFFF", bg=COLOR_HEADER_SLATE)

    sc = {"UP": (COLOR_UP_GREEN, COLOR_UP_TXT), "DOWN": (COLOR_DOWN_RED, COLOR_DOWN_TXT), "UNASSIGNED": (COLOR_UNASG_GREY, COLOR_UNASG_TXT)}

    for idx, r in enumerate(results, 1):
        device, hostname, ip, vlan, loc, status, avg_ms, last = r
        rn = 10 + idx
        ws.row_dimensions[rn].height = 18
        bg2 = "FFFFFF" if idx % 2 == 0 else "F8F9FA"
        sbg, sfc = sc.get(status, ("FFFFFF", "000000"))
        pkt = "0%" if status == "UP" else "100%"
        ms_v = f"{avg_ms:.1f} ms" if avg_ms else "-"
        ts = now.strftime("%d-%b %H:%M")
        daily_uptime, daily_downtime = get_daily_stats(ip)
        rem = ("Reachable/Nominal" if status == "UP" else "No response/Outage Risk" if status == "DOWN" else "Unassigned System IP")
        if avg_ms and avg_ms > HIGH_LAT_MS:
            rem = "High Latency Degraded"

        row_vals = [idx, ip, device, hostname, vlan, loc, status, ms_v, pkt, ts,daily_uptime,daily_downtime,last, rem]
        for ci, val in enumerate(row_vals, 2):
            if ci == 8:  # Status column style block
                _cell(ws, rn, ci, val, bold=True, size=8, fc=sfc, bg=sbg)
            elif ci == 9:  # Precise Latency Tracking Styles
                if avg_ms:
                    ms_bg2, ms_fc2 = ms_color(avg_ms)
                    _cell(ws, rn, ci, val, bold=True, size=8, fc=ms_fc2, bg=ms_bg2)
                else:
                    _cell(ws, rn, ci, val, size=8, bg=bg2)
            elif ci == 13 and rem == "High Latency Degraded":
                _cell(ws, rn, ci, val, size=8, bg=COLOR_WARN_YLLW, fc=COLOR_WARN_TXT, bold=True)
            elif ci in (11, 12, 13):
                _cell(ws, rn, ci, val, size=8, bg=bg2, halign="left", wrap=True)
            else:
                _cell(ws, rn, ci, val, size=8, bg=bg2)

# ══════════════════════════════════════════════    
#  EMAIL SYSTEMS & WORKSPACE RUNNER
# ══════════════════════════════════════════════
def dispatch_management_alert(path, summary_dict, all_results, scan_duration):
    if not GMAIL_SENDER or not MANAGER_EMAIL:
        return
    now = datetime.now()
    date_str = now.strftime("%d-%b-%Y")
    time_str = now.strftime("%I:%M %p")
    os_info = f"{platform.system()} {platform.release()}"

    msg = MIMEMultipart()
    msg["From"] = GMAIL_SENDER
    msg["To"]   = ", ".join(MANAGER_EMAIL)
    msg["Subject"] = f"SNOC Network Ping Monitoring Report - {date_str}"

    # Calculate overall stats
    grand_t = grand_u = grand_d = grand_un = 0
    for sname, res in all_results.items():
        grand_t += len(res)
        grand_u += sum(1 for r in res if r[5] == "UP")
        grand_d += sum(1 for r in res if r[5] == "DOWN")
        grand_un += sum(1 for r in res if r[5] == "UNASSIGNED")
        
    overall_avail = (grand_u / grand_t * 100) if grand_t else 0

    # Helper for Grade logic matching your main sheet
    def get_grade_info(avail):
        if avail >= 95: return "A", "EXCELLENT STATUS"
        if avail >= 85: return "B", "MOSTLY OPERATIONAL"
        if avail >= 75: return "C", "ATTENTION REQUIRED"
        return "D", "CRITICAL INCIDENT"

    overall_grade, overall_phrase = get_grade_info(overall_avail)

    # 1. Alert HTML (Only shows if there are down nodes)
    alert_html = ""
    if grand_d > 0:
        alert_html = f"""
        <div style="background-color: #FEF2F2; border-left: 6px solid #C53030; padding: 16px; margin: 20px 0; border-radius: 4px;">
            <span style="color: #C53030; font-size: 15px; font-weight: bold;">⚠️ ALERT: {grand_d} device(s) are UNREACHABLE. Immediate attention required.</span>
        </div>
        """

    # 2. Build Category Table Rows dynamically
    table_rows = ""
    for sname, res in all_results.items():
        t = len(res)
        u = sum(1 for r in res if r[5] == "UP")
        d = sum(1 for r in res if r[5] == "DOWN")
        un = sum(1 for r in res if r[5] == "UNASSIGNED")
        av = (u / t * 100) if t else 0
        gr, _ = get_grade_info(av)
        
        status_dot = "🟢" if d == 0 else "🔴"
        row_bg = "#FFFFFF" if d == 0 else "#FFF5F5"
        
        # Color specific formatting for table columns
        u_color = "#16A34A"
        d_color = "#DC2626" if d > 0 else "#16A34A"
        gr_color = "#16A34A" if gr == "A" else "#2563EB" if gr == "B" else "#D97706" if gr == "C" else "#DC2626"

        table_rows += f"""
        <tr style="background-color: {row_bg}; font-size: 13px; text-align: center; border-bottom: 1px solid #E5E7EB;">
            <td style="padding: 10px; border-bottom: 1px solid #E5E7EB;">{status_dot}</td>
            <td style="padding: 10px; text-align: left; font-weight: 600; color: #374151; border-bottom: 1px solid #E5E7EB;">{sname}</td>
            <td style="padding: 10px; color: {u_color}; font-weight: bold; border-bottom: 1px solid #E5E7EB;">{u}</td>
            <td style="padding: 10px; color: {d_color}; font-weight: bold; border-bottom: 1px solid #E5E7EB;">{d}</td>
            <td style="padding: 10px; color: #6B7280; border-bottom: 1px solid #E5E7EB;">{un}</td>
            <td style="padding: 10px; font-weight: 600; color: #374151; border-bottom: 1px solid #E5E7EB;">{av:.1f}%</td>
            <td style="padding: 10px; color: {gr_color}; font-weight: bold; border-bottom: 1px solid #E5E7EB;">{gr}</td>
        </tr>
        """

    # 3. Main Body HTML
    body = f"""
    <html>
    <head>
        <meta charset="utf-8">
    </head>
    <body style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #F3F4F6; margin: 0; padding: 20px;">
        
        <table align="center" width="800" cellpadding="0" cellspacing="0" style="background-color: #FFFFFF; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); overflow: hidden;">
            
            <!-- HEADER LOGO / TITLE -->
            <tr>
                <td align="center" style="background-color: #1B2A40; padding: 25px;">
                    <h1 style="margin: 0; font-size: 20px; font-weight: 600; color: #FFFFFF; letter-spacing: 0.5px;">📡 SNOC NETWORK PING MONITORING REPORT</h1>
                    <p style="margin: 8px 0 0; font-size: 12px; color: #94A3B8;">Karnataka | Bangalore | India</p>
                </td>
            </tr>

            <!-- INFO BAR -->
            <tr>
                <td align="center" style="background-color: #226089; padding: 10px 25px; color: #FFFFFF; font-size: 12px; font-weight: 600;">
                    📅 {date_str} &nbsp;|&nbsp; ⏱️ {time_str} &nbsp;|&nbsp; 💻 {os_info} &nbsp;|&nbsp; 
                </td>
            </tr>

            <!-- MAIN CONTENT -->
            <tr>
                <td style="padding: 30px;">
                    
                    <p style="margin: 0 0 15px; font-size: 14px; color: #374151;">Dear Team,</p>
                    <p style="margin: 0 0 20px; font-size: 14px; color: #4B5563; line-height: 1.6;">
                        Please find the attached <b>SNOC Network Ping Monitoring Report</b> generated on <b>{date_str} at {time_str}</b>. The Excel file contains a detailed per-category report with full ping results, response times, and availability statistics.
                    </p>

                    <!-- CRITICAL ALERT INJECTION -->
                    {alert_html}

                    <!-- KPI METRICS ROW -->
                    <table width="100%" cellpadding="0" cellspacing="0" style="margin: 25px 0;">
                        <tr>
                            <td width="23%" align="center" style="background-color: #E4EEF5; padding: 20px; border-radius: 6px;">
                                <span style="font-size: 11px; font-weight: 700; color: #1E40AF;">TOTAL IPs</span><br>
                                <span style="font-size: 32px; font-weight: 700; color: #1E3A8A; display: block; margin-top: 8px;">{grand_t}</span>
                            </td>
                            <td width="2.6%"></td>
                            <td width="23%" align="center" style="background-color: #E8F5E9; padding: 20px; border-radius: 6px;">
                                <span style="font-size: 11px; font-weight: 700; color: #166534;">✅ REACHABLE</span><br>
                                <span style="font-size: 32px; font-weight: 700; color: #14532D; display: block; margin-top: 8px;">{grand_u}</span>
                            </td>
                            <td width="2.6%"></td>
                            <td width="23%" align="center" style="background-color: #FEF2F2; padding: 20px; border-radius: 6px;">
                                <span style="font-size: 11px; font-weight: 700; color: #991B1B;">❌ DOWN</span><br>
                                <span style="font-size: 32px; font-weight: 700; color: #7F1D1D; display: block; margin-top: 8px;">{grand_d}</span>
                            </td>
                            <td width="2.6%"></td>
                            <td width="23%" align="center" style="background-color: #FEF3C7; padding: 20px; border-radius: 6px;">
                                <span style="font-size: 11px; font-weight: 700; color: #92400E;">📊 AVAILABILITY &nbsp;<span style="color: #451A03;">GRADE: {overall_grade}</span></span><br>
                                <span style="font-size: 28px; font-weight: 700; color: #78350F; display: block; margin-top: 8px;">{overall_avail:.1f}%</span>
                                <span style="font-size: 10px; font-weight: 600; color: #92400E; display: block; margin-top: 4px; text-transform: uppercase;">{overall_phrase}</span>
                            </td>
                        </tr>
                    </table>

                    <!-- TABLE HEADER TITLE -->
                    <h3 style="font-size: 15px; color: #1F2937; margin: 30px 0 10px; border-bottom: 2px solid #1F2937; padding-bottom: 8px;">
                        📋 PER-CATEGORY SUMMARY
                    </h3>

                    <!-- DATA TABLE -->
                    <table width="100%" cellpadding="0" cellspacing="0" style="border: 1px solid #E5E7EB; border-collapse: collapse;">
                        <thead>
                            <tr style="background-color: #226089; color: #FFFFFF; font-size: 12px; text-align: center;">
                                <th style="padding: 12px; width: 5%;"></th>
                                <th style="padding: 12px; text-align: left;">Category</th>
                                <th style="padding: 12px;">UP ✅</th>
                                <th style="padding: 12px;">DOWN ❌</th>
                                <th style="padding: 12px;">UNASSIGNED</th>
                                <th style="padding: 12px;">AVAILABILITY</th>
                                <th style="padding: 12px;">GRADE</th>
                            </tr>
                        </thead>
                        <tbody>
                            {table_rows}
                        </tbody>
                    </table>

                    <p style="margin: 25px 0 0; font-size: 13px; color: #4B5563;">
                        Please refer to the <b>attached Excel report</b> for complete details including response times, high latency alerts, and per-device status.
                    </p>

                </td>
            </tr>

            <!-- FOOTER -->
            <tr>
                <td align="center" style="background-color: #1B2A40; padding: 15px; color: #94A3B8; font-size: 11px;">
                    Auto-generated by <b>SNOC Network Monitoring System</b> &nbsp;|&nbsp; Do not reply to this email &nbsp;|&nbsp; Contact SNOC Network Team for issues
                </td>
            </tr>

        </table>

    </body>
    </html>
    """
    
    msg.attach(MIMEText(body, "html"))
    
    # 4. Attachment & Sending execution
    try:
        with open(path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(path)}")
        msg.attach(part)
        
        cfg = {"sender": GMAIL_SENDER, "password": GMAIL_PASSWORD, "manager": MANAGER_EMAIL}
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(cfg["sender"], cfg["password"])
            s.sendmail(cfg["sender"], cfg["manager"], msg.as_string())
        print(f"  [✓] Management Executive Alert Email sent to target vector: {cfg['manager']}")
    except Exception as e:
        print(f"  [❌] Critical: Failed to route management email vector line: {e}")

import pandas as pd
#=====================================================================
# thsi function for uptime calculation 
#=====================================================================
def update_uptime_history(all_results):
    now = datetime.now()

    rows = []

    for sname, results in all_results.items():
        for device, hostname, ip, vlan, location, status, avg_ms, last in results:
            rows.append({
                "Timestamp": now,
                "IP": ip,
                "Status": status
            })

    df_new = pd.DataFrame(rows)

    if os.path.exists(HISTORY_FILE):
        df_old = pd.read_excel(HISTORY_FILE)
        df = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df = df_new

    df.to_excel(HISTORY_FILE, index=False)
def get_daily_stats(ip):
    if not os.path.exists(HISTORY_FILE):
        return "100.00%", "0 Hr"

    df = pd.read_excel(HISTORY_FILE)

    today = datetime.now().date()

    df["Timestamp"] = pd.to_datetime(df["Timestamp"])

    df = df[
        (df["IP"] == ip)
        &
        (df["Timestamp"].dt.date == today)
    ]

    total = len(df)

    if total == 0:
        return "100.00%", "0 Hr"

    up_count = len(df[df["Status"]=="UP"])
    down_count = len(df[df["Status"]=="DOWN"])

    uptime = up_count/total*100
    downtime_hours = down_count * CHECK_INTERVAL_HOURS

    return f"{uptime:.2f}%", f"{downtime_hours} Hr"

#==================================================
#SLA REPORT SHEET
# ================================================= 
def build_sla_report(ws):
    ws.append([
        "IP Address",
        "Weekly Uptime %",
        "Weekly Downtime",
        "Monthly Uptime %",
        "Monthly Downtime"
    ])
    #========================================
def run_job():
    now = datetime.now()

    print(f"\n[{now.strftime('%Y-%m-%d %H:%M:%S')}] Starting target validation matrix...")

    if not os.path.exists(INPUT_EXCEL):
        print(f"  [❌] Critical Exception: Base structural matrix sheet '{INPUT_EXCEL}' is missing.")
        return

    start_time = time.time()

    try:
        sheet_data = extract_sheet_ips(INPUT_EXCEL)
    except Exception as e:
        print(f"  [❌] Data Isolation Fault reading '{INPUT_EXCEL}': {e}")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    exec_ws = wb.create_sheet("⭐ EXECUTIVE SUMMARY")
    dash_ws = wb.create_sheet("📊 SYSTEM DASHBOARD")
    sla_ws  = wb.create_sheet("📈 SLA REPORT")

    all_results = {}
    summary = {}

    # =========================================
    # Ping all IPs and collect results
    # =========================================
    for sname, ip_tuples in sheet_data.items():

        print(f"  -> Processing cluster sector: [{sname}] ({len(ip_tuples)} nodes)...")

        results = []

        for device, hostname, ip, vlan, location in ip_tuples:

            status, avg_ms, last = ping(ip)

            print(f"     {ip:<16} -> {status:<12}")

            results.append(
                (
                    device,
                    hostname,
                    ip,
                    vlan,
                    location,
                    status,
                    avg_ms,
                    last
                )
            )

        all_results[sname] = results

        u = sum(1 for r in results if r[5] == "UP")
        d = sum(1 for r in results if r[5] == "DOWN")
        un = sum(1 for r in results if r[5] == "UNASSIGNED")

        summary[sname] = (u, d, un)

    # =========================================
    # Update history ONCE
    # =========================================
    update_uptime_history(all_results)

    # =========================================
    # Build device sheets
    # =========================================
    for sname, results in all_results.items():

        ws = wb.create_sheet(sname[:31])

        build_sheet_report(
            ws,
            sname,
            results,
            now
        )

    elapsed = time.time() - start_time
    scan_duration = f"{int(elapsed//60)}m {int(elapsed%60)}s"

    # =========================================
    # Build summary sheets
    # =========================================
    build_executive_summary(exec_ws, all_results, now, scan_duration)

    build_dashboard(dash_ws, all_results, now)

    build_sla_report(sla_ws)

    # =========================================
    # Save report
    # =========================================
    os.makedirs(REPORT_FOLDER, exist_ok=True)

    ts = now.strftime("%Y-%m-%d_%H-%M")

    out = os.path.join(
        REPORT_FOLDER,
        f"SNOC_Ping_Report_{ts}.xlsx"
    )

    wb.save(out)

    print(f"\n[✓] Clean Management Report Generated -> {out}")

    print(f"[✓] Execution Phase Complete. Span: {scan_duration}")

    dispatch_management_alert(
        out,
        summary,
        all_results,
        scan_duration
    )

if __name__ == "__main__":
    run_job()

#if __name__ == "__main__":
 #   print("═" * 60)
  #  print("     SNOC PREMIUM IP MONITORING & ENTERPRISE METRICS ENGINE")
   # print("═" * 60)
    #run_job()
    #print("\n[INFO] Setting engine execution timeline schedule to check every 2 Hours...")
    #schedule.every(2).hours.do(run_job)
    #while True:
     #   schedule.run_pending()
      #  time.sleep(1)